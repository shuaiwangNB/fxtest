import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Fill, Border, Side, PatternFill
from openpyxl.utils import quote_sheetname
from fxtest.logging import log
import os

def create_excel(path, rows=1000, sheet_name="Sheet1"):

    if len(path)>3:
        if path[-5:]==".xlsx":
            excel_path=path
        else:
            log.error("文件格式不正确")
            return 0
    else:
        log.error("文件不正确")
        return 0

    wb = Workbook()
    # wb = load_workbook(path)
    wb.active
    sheet = wb.create_sheet(sheet_name, 0)
    sheet2 = wb.create_sheet("系统数据", 1)
    # sheet = wb["Sheet1"]
    titles = ["用例标题", "页面操作", "元素路径", "输入内容", "其他说明"]
    types = ["get", "input_text", "assertElementisExist", "assertText",
             "assertElementText", "wait", "wait_elements", "refresh", "click", "slow_click"]
    # sheet2=wb["Sheet2"]

    for k, t in enumerate(types):
        sheet2["A{}".format(k+1)].value = t

    border = Border(left=Side(border_style="thin", color="00000000"),
                    right=Side(border_style="thin", color="00000000"),
                    top=Side(border_style="thin", color="00000000"),
                    bottom=Side(border_style="thin", color="00000000")
                    )
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(name="等线", size=11, bold=False, italic=False, color='00FFFFFF')
    for k, title in enumerate(titles):
        col = chr(ord("A")+k)+"1"
        cell = sheet[col]
        cell.border = border
        cell.fill = fill
        cell.font = font
        cell.value = title
        sheet.column_dimensions[chr(ord("A")+k)].width = 22
    sheet.row_dimensions[1].height = 22

    dv = DataValidation(type="list", formula1="{0}!$A$1:$A${1}".format(
        quote_sheetname("系统数据"), len(types)), allow_blank=True)
    sheet.add_data_validation(dv)
    row = 2
    for _ in range(rows):
        dv.add(sheet["B{}".format(row)])
        row += 1
    wb.save(excel_path)
