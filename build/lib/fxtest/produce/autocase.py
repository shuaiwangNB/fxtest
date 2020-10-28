from openpyxl import Workbook, load_workbook
import os
from fxtest.logging import log

def code_write(case: list):
    try:
        if case[1] == "get":
            code = "self.{0}(\"{1}\")".format(case[1], case[3])
        if case[1] == "input_text":
            if case[4] != None:
                code = "self.{0}(text='{1}',xpath=\"{2}\",{3})".format(
                    case[1], case[3], case[2], case[4])
            else:
                code = "self.{0}(text='{1}',xpath=\"{2}\")".format(
                    case[1], case[3], case[2])
        if case[1] == "assertElementisExist":
            code = "self.{0}(xpath=\"{1}\",msg=\"{2}\")".format(
                case[1], case[2], case[4])
        if case[1] == "assertText":
            code = "self.{0}(text='{1}',msg=\"{2}\")".format(
                case[1], case[3], case[4])
        if case[1] == "assertElementText":
            code = "self.{0}(text='{1}',xpath=\"{2}\",msg=\"{3}\")".format(
                case[1], case[3], case[2], case[4])
        if case[1] == "wait":
            code = "self.{0}(time=\"{1}\")".format(case[1], case[3])
        if case[1] == "wait_elements":
            code = "self.{0}(element=\"{1}\")".format(case[1], case[2])
        if case[1] == "refresh":
            code = "self.{0}():".format(case[1])
        if case[1] == "click":
            code = "self.{0}(xpath=\"{1}\")".format(case[1], case[2])
        if case[1] == "slow_click":
            code = "self.{0}(xpath=\"{1}\")".format(case[1], case[2])
        # else:
        #     raise IndexError("excel case is nonstandard")

        return code
    except IndexError:
        print("excel case is nonstandard")


def create_test_case(path, py_path,sheet_name="Sheet1"):
    if len(path)>3:
        if os.path.exists(path):

            if path[-5:]==".xlsx":
                excel_path=path
            else:
                log.error("文件格式不正确")
                return 0
        else:
            log.error("文件不存在")
            return 0
    else:
        log.error("文件不正确")
        return 0
    
    if len(py_path)>3:
        if os.path.exists(py_path) is False:
            if path[-3:]==".py":
                py_path=py_path
            else:
                log.error("文件格式不正确")
                return 0
        else:
            log.error("文件已存在")
            return 0
    else:
        log.error("文件不正确")
        return 0
    test_sample = """
import fxtest
import pytest

class Test_{class_name}(fxtest.TestCase):
"""
    class_name = os.path.basename(excel_path).split(".")[0]
    test_sample = test_sample.format(class_name=class_name)
    wb = load_workbook(excel_path)
    flag = None
    sheet = wb.get_sheet_by_name(sheet_name)
    col_range = sheet["A"]

    case_intervel = []

    for k, col in enumerate(col_range):
        if col.value != None:
            if col.value != "用例标题":
                case_intervel.append(k)

    for k, case in enumerate(case_intervel):

        while True:
            if k < len(case_intervel)-1:

                if case == case_intervel[k+1]:
                    break
            else:
                if case == sheet.max_row:
                    break
            case_content = []
            for field in list(sheet.rows)[case]:
                case_content.append(field.value)
            if case_content[0] != None:
                try:
                    test_sample = test_sample+"\r"+" "*4 + \
                        "def test_{}(self):".format(case_content[0])
                    test_sample = test_sample+"\r" + \
                        " "*8+code_write(case_content)
                except Exception:
                    return

            else:
                code = code_write(case_content)
                test_sample = test_sample+"\r"+" "*8+code
            case += 1

    test_end_sample = """
if __name__ =="__main__":
    fxtest.main()
"""

    with open("D:\\fxyb\\test.py", "w+", encoding="utf-8") as f:
        f.write(test_sample+"\r"+test_end_sample)


if __name__ == "__main__":
    pass