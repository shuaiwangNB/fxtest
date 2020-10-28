import openpyxl
from openpyxl import load_workbook
import yaml



def yaml_to_list(file, key=None,s_key=None):
    """
    """
    if file is None:
        raise FileExistsError("Please specify the YAML file to convert.")

    if key is None:
        with open(file, "r", encoding="utf-8") as f:
            dict_data = yaml.safe_load(f)
    else:
        with open(file, "r", encoding="utf-8") as f:
            try:
                if s_key and key != None:
                    dict_data=yaml.safe_load(f)[key][s_key]
                else:
                    dict_data = yaml.safe_load(f)[key]
            except KeyError:
                raise ValueError("Check the test data, no '{}'".format(key))

    return dict_data

def excel_to_list(file=None, sheet="Sheet1", line=1):
    """

    """
    if file is None:
        raise FileExistsError("Please specify the Excel file to convert.")

    excel_table = load_workbook(file)
    sheet = excel_table[sheet]

    table_data = []
    for line in sheet.iter_rows(line, sheet.max_row):
        line_data = []
        for field in line:
            line_data.append(field.value)
        table_data.append(line_data)

    return table_data